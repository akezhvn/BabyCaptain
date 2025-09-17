import telebot
from telebot import types
import openpyxl

bot = telebot.TeleBot("SECRET")

user_data = {}


# excel
def save_data(fio, phone, ticket_number, user_id):
    wb = openpyxl.load_workbook('data.xlsx')
    ws = wb.active
    ws.append([fio, phone, ticket_number, user_id])
    wb.save('data.xlsx')


# handler for start
@bot.message_handler(commands=['start'])
def start_message(message):
    bot.send_message(message.chat.id, "Сәлем! Аты-жөніңізді енгізіңіз:")
    user_data[message.chat.id] = {}
    bot.register_next_step_handler(message, input_fio)


# name
def input_fio(message):
    user_data[message.chat.id]['fio'] = message.text
    bot.send_message(message.chat.id, "Введите номер телефона:")
    bot.register_next_step_handler(message, input_phone)


# phone
def input_phone(message):
    user_data[message.chat.id]['phone'] = message.text
    bot.send_message(message.chat.id, "Билет номерін енгізіңіз:")
    bot.register_next_step_handler(message, input_ticket_number)


#ticket number
def input_ticket_number(message):
    if ticket_number_exists(message.text):
        bot.send_message(message.chat.id, "Бұл билет тіркелген, басқа номер енгізіңіз::")
        bot.register_next_step_handler(message, input_ticket_number)
    else:
        user_data[message.chat.id]['ticket_number'] = message.text
        bot.send_message(message.chat.id, "Билеттің суретін жіберіңіз:")
        bot.register_next_step_handler(message, input_ticket_photo)


# ticket image
def input_ticket_photo(message):
    user_id = message.from_user.id
    ticket_number = user_data[message.chat.id]['ticket_number']
    if message.photo:
        file_id = message.photo[-1].file_id
        file_info = bot.get_file(file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        file_path = "photos/{}.jpg".format(ticket_number)
        with open(file_path, 'wb') as new_file:
            new_file.write(downloaded_file)
        save_data(user_data[message.chat.id]['fio'], user_data[message.chat.id]['phone'], ticket_number, user_id)
        bot.send_message(message.chat.id, "Құттықтаймын! Сіздің билетіңіз тіркелді!♥")
        bot.send_message(message.chat.id, "Жаңа билет тіркеу үшін /start батырмасын басыңыз.")
    else:
        bot.send_message(message.chat.id, "Сурет жіберіңіз өтініш!")
        bot.register_next_step_handler(message, input_ticket_photo)


def ticket_number_exists(ticket_number):
    wb = openpyxl.load_workbook('data.xlsx')
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if row[2] == ticket_number:
            return True
    return False



bot.polling()
